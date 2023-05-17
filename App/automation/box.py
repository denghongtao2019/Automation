import logging
import smtplib
import sys
import unittest
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os

import openpyxl
import yaml
import pymysql
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select


class BoxDriver:
    def __init__(self, browser):
        '''
        实例化浏览器对象的构造方法
        :param browser: 浏览器名称
        '''
        try:
            if browser == 'Chrome':
                driver = webdriver.Chrome()
            elif browser == 'Firefox':
                driver = webdriver.Firefox()
            else:
                driver = webdriver.Ie()
            # 把driver转换为私有的全局变量
            self._base_driver = driver
        except:
            raise NameError('请输入有效的浏览器名称(Chrome/Firefox/Ie)')

    def _convert_location(self, selector):
        '''
        元素定位的初步处理，把自定义的selector元素定位的方式和值，转换为selenium标准的定位格式
        :param selector: 元素定位的方式和值，如：'id,account'
        :return: selenium标准的定位格式，如：(By.ID, 'account')
        '''
        location_by = selector.split(',')[0]
        location_value = selector.split(',')[1]
        if location_by == 'i' or location_by == 'id':
            location = (By.ID, location_value)
        elif location_by == 'cl' or location_by == 'class_name':
            location = (By.CLASS_NAME, location_value)
        elif location_by == 'n' or location_by == 'name':
            location = (By.NAME, location_value)
        elif location_by == 'l' or location_by == 'link_text':
            location = (By.LINK_TEXT, location_value)
        elif location_by == 'pl' or location_by == 'partial_link_text':
            location = (By.PARTIAL_LINK_TEXT, location_value)
        elif location_by == 'x' or location_by == 'xpath':
            location = (By.XPATH, location_value)
        elif location_by == 'css' or location_by == 'css_selector':
            location = (By.CSS_SELECTOR, location_value)
        elif location_by == 'tag' or location_by == 'tag_name':
            location = (By.TAG_NAME, location_value)
        else:
            raise NameError('请输入一个正确的定位方式')
        return location

    def _get_element(self, selector):
        '''
        定位单个元素
        :param selector: 元素定位的方式和值，如：'id,account'
        :return: 定位的目标元素
        '''
        location = self._convert_location(selector)
        ele = self._base_driver.find_element(*location)
        return ele

    def _get_elements(self, selector):
        '''
        定位一组元素
        :param selector: 元素定位的方式和值，如：'id,account'
        :return: 定位的目标元素组
        '''
        location = self._convert_location(selector)
        eles = self._base_driver.find_elements(*location)
        return eles

    def get(self, url):
        '''
        打开url对应的网站
        :param url: 网站地址
        :return:
        '''
        self._base_driver.get(url)

    def maximize_window(self):
        '''浏览器窗口最大化'''
        self._base_driver.maximize_window()

    def send_keys(self, selector, text):
        '''
        写入输入框内容
        :param selector: 元素定位的方式和值，如：'id,account'
        :param text: 输入的文本
        :return:
        '''
        ele = self._get_element(selector)
        ele.clear()
        ele.send_keys(text)

    def click(self, selector):
        '''
        对元素进行点击操作
        :param selector: 元素定位的方式和值，如：'id,account'
        :return:
        '''
        ele = self._get_element(selector)
        ele.click()

    def switch_to_frame(self, id):
        '''
        切换frame
        :param id: frame的id值
        :return:
        '''
        self._base_driver.switch_to.frame(id)

    def get_element_text(self, selector):
        '''
        获取一个元素的文本
        :param selector: 元素定位方式和值，如'id,account'
        :return:一个元素的文本
        '''
        ele = self._get_element(selector)
        text = ele.text
        return text

    def get_elements_text(self, selector, index):
        '''
        获取一组元素中的一个元素的文本
        :param selector: 元素定位方式和值，如'id,account'
        :param index: 元素的索引值
        :return:一个元素的文本
        '''
        eles = self._get_elements(selector)
        text = eles[index].text
        return text

    def select_by_choice(self, selector, choice, content):
        '''
        下拉框选择内容
        :param selector: 元素定位方式和值，如'id,account'
        :param select:选择方式
        :param content:选择方式对应的值
        :return:
        '''
        # 1、定位下拉框
        ele = self._get_element(selector)
        # 2、引入Select类对下拉框操作
        select_dept = Select(ele)
        # 3、对下拉框内容进行选择
        if choice == 't' or choice == 'text':
            select_dept.select_by_visible_text(content)
        elif choice == 'v' or choice == 'value':
            select_dept.select_by_value(content)
        elif choice == 'i' or choice == 'index':
            select_dept.select_by_index(content)
        else:
            raise NameError('请输入有效的下拉选择方式！')

    def select_by_visible_text(self, selector, visible_text):
        '''
        visible_text方式下拉选择内容
        :param selector: 元素定位方式和值，如'id,account'
        :param visible_text: 下拉内容的visible_text值
        :return:
        '''
        # 1、定位下拉框
        ele = self._get_element(selector)
        # 2、引入Select类对下拉框操作
        select_dept = Select(ele)
        # 3、对下拉框内容进行选择
        select_dept.select_by_visible_text(visible_text)

    def select_by_value(self, selector, value):
        '''
        value方式下拉选择内容
        :param selector: 元素定位方式和值，如'id,account'
        :param value: 下拉内容的value值
        :return:
        '''
        # 1、定位下拉框
        ele = self._get_element(selector)
        # 2、引入Select类对下拉框操作
        select_dept = Select(ele)
        # 3、对下拉框内容进行选择
        select_dept.select_by_value(value)

    def select_by_index(self, selector, index):
        '''
        index方式下拉选择内容
        :param selector: 元素定位方式和值，如'id,account'
        :param index: 下拉框内容的索引值
        :return:
        '''
        # 1、定位下拉框
        ele = self._get_element(selector)
        # 2、引入Select类对下拉框操作
        select_dept = Select(ele)
        # 3、对下拉框内容进行选择
        select_dept.select_by_index(index)

    def implicitly_wait(self, second):
        '''
        隐式等待
        :param second: 等待的时间
        :return:
        '''
        self._base_driver.implicitly_wait(second)

    def close(self):
        '''关闭当前浏览器'''
        self._base_driver.close()

    def quit(self):
        '''退出所有浏览器'''
        self._base_driver.quit()

    def get_screenshot_as_file(self, filename):
        '''
        截图
        :param filename: 截图后生成文件的路径
        :return:
        '''
        self._base_driver.get_screenshot_as_file(filename)


class BaseDriver:
    def __init__(self, driver: BoxDriver):
        '''
        用于继承的父类，指定driver的值
        :param driver: （driver:BoxDriver）这种格式是指定的driver参数，必须只能是BoxDriver类实例化的对象
        '''
        self.base_driver = driver


class DbHelper:

    def __init__(self, host, port, username, pwd, db_name):
        '''
        建立数据库连接和游标
        :param host: 数据库主机ip
        :param port: 数据库端口，需要输入整型，如3306
        :param username: 数据库登录的用户名
        :param pwd: 数据库登录的密码
        :param db_name: 数据库名称
        :return:
        '''
        # 建立数据库连接
        self.db_connect = pymysql.connect(
            host=host, port=port, user=username, password=pwd, database=db_name, charset='utf8')
        # 建立游标
        self.get_cursor = self.db_connect.cursor()

    def execute_sql(self, sql):
        '''
        执行sql语句
        :param sql: sql语句
        :return:
        '''
        self.get_cursor.execute(sql)

    def get_db_data(self, sql):
        '''
        返回所有的查询数据
        :return: 所有的查询数据
        '''
        self.get_cursor.execute(sql)
        all_data = self.get_cursor.fetchall()
        return all_data

    def close_db(self):
        '''关闭数据库连接'''
        self.get_cursor.close()
        self.db_connect.close()


class YamlHelper:
    def get_yaml_data(self, yaml_file):
        '''
        读取yaml文件数据
        :param yaml_file: yaml数据文件名称
        格式：{'Person': {'name': 'xiaoming', 'age': '18'}, 'animal': {'dog': 'wangwang', 'cat': 'miaomiao'}}
        :return: yaml文件数据，格式，
        '''
        with open(yaml_file, mode='r', encoding='utf8') as yaml_file:
            yaml_data = yaml.safe_load(yaml_file)
            return yaml_data


class ExcelHelper:
    def get_excel_data(self, excel_file, sheet_name):
        # 获取excel文件
        get_excle = openpyxl.load_workbook(excel_file)
        # 获取sheet页
        get_sheet = get_excle[sheet_name]
        # 定义空列表，用来保存每行的数据
        get_one_line_list = []
        # 定义空列表，用来保存所有行的数据
        get_all_line_list = []
        par = True
        for get_line in get_sheet:
            # 过滤标题
            if par:
                par = False
                continue
            for get_cell in get_line:
                if get_cell.value == None:
                    get_cell.value = ''
                # 获取每个单元格的数据，一行获取多个，如：xiaoming1，小明
                get_one_line_list.append(str(get_cell.value))
                # 获取完每一行数据后，把数据转换为元组
            get_one_line_tuple = tuple(get_one_line_list)
            # 获取完每一行数据后，把转换为元组的数据追加到一个列表中
            get_all_line_list.append(get_one_line_tuple)
            # 清空列表，避免重复
            get_one_line_list.clear()
        return get_all_line_list

    def write_excel_data(self, excel_file, sheet_name, row, column, result):
        # 打开excel文件
        get_excle = openpyxl.load_workbook(excel_file)
        # 获取sheet页
        get_sheet = get_excle[sheet_name]
        # 写入数据
        get_sheet.cell(row, column).value = result
        # 保存关闭文件
        get_excle.save(excel_file)


class AssertHelper(unittest.TestCase):
    '''断言方法'''

    def assert_equal(self, actValue, expValue):
        try:
            self.assertEqual(actValue, expValue)
        except Exception as err:
            assert False, "\n" \
                          "实际结果和预期结果不一致，实际结果为:%s,预期结果为:%s" % (actValue, expValue)

    def assert_not_equal(self, actValue, expValue):
        try:
            self.assertNotEqual(actValue, expValue)
        except Exception as err:
            assert False, "\n" \
                          "实际结果和预期结果一致，实际结果为:%s,预期结果为:%s" % (actValue, expValue)

    def assert_in(self, actValue, expValue):
        try:
            self.assertIn(actValue, expValue)
        except Exception as err:
            assert False, "\n" \
                          "实际结果不在预期结果中，实际结果为:%s,预期结果为:%s" % (actValue, expValue)

    def assert_not_in(self, actValue, expValue):
        try:
            self.assertNotIn(actValue, expValue)
        except Exception as err:
            assert False, "\n" \
                          "实际结果在预期结果中，实际结果为:%s,预期结果为:%s" % (actValue, expValue)

    def assert_ture(self, result):
        try:
            self.assertTrue(result)
        except Exception as err:
            assert False, "结果为:%s" % result

    def assert_false(self, result):
        try:
            self.assertFalse(result)
        except Exception as err:
            assert False, "结果为:%s" % result

    def assert_is_none(self, obj):
        try:
            self.assertIsNone(obj)
        except Exception as err:
            assert False, "结果不为空"

    def assert_is_not_none(self, obj):
        try:
            self.assertIsNotNone(obj)
        except Exception as err:
            assert False, "结果为空"


class LogHelper:
    def __init__(self, log_path):
        # 定义一个变量，存储日志路径,带上日志文件的名称
        self.file_name = log_path
        # 通过logging模块下的getLogger方法，得到一个可操作日志对象loger
        self.loger = logging.getLogger()
        # 设置日志的级别
        self.loger.setLevel(logging.DEBUG)
        # 设置日志的格式和内容
        self.formatter = logging.Formatter(
            '[%(asctime)s]-[%(filename)s]-[%(levelname)s]:[%(message)s]')

    def _console(self, level, message):
        '''
        日志模块的核心处理功能
        :param level: 日志等级
        :param message: 日志消息数据
        :return:
        '''
        # 创建一个FileHandler对象，将日志写入到文件，把日志内容追加到文件的末尾
        fh = logging.FileHandler(self.file_name, 'a', 'utf8')
        fh.setLevel(logging.DEBUG)
        # 设置文件内容的格式
        fh.setFormatter(self.formatter)
        # 将内容添加到日志文件
        self.loger.addHandler(fh)
        # 创建一个StreamHandler对象，把日志输出到控制台
        ch = logging.StreamHandler(sys.stdout)
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.loger.addHandler(ch)

        if level == 'info':
            self.loger.info(message)
        elif level == 'debug':
            self.loger.debug(message)
        elif level == 'warning':
            self.loger.warning(message)
        elif level == 'error':
            self.loger.error(message)

        # 为了避免日志的重复输入
        self.loger.removeHandler(fh)
        self.loger.removeHandler(ch)
        fh.close()

    def info(self, message):
        self._console('info', message)

    def debug(self, message):
        self._console('debug', message)

    def warning(self, message):
        self._console('warning', message)

    def error(self, message):
        self._console('error', message)


class EmailHelper:

    def email_send_163(self, report_file, mail_title, email_receiver):
        '''发送163邮箱的方法'''
        '''配置邮件服务器信息'''
        '''发件相关参数'''
        try:
            # 发件服务器
            smtpserver = 'smtp.163.com'
            # 端口号
            port = 25
            # 发件人邮箱
            sender = ''
            # 授权码
            pwd = ''
            # 收件人邮箱
            receiver = email_receiver
            # 创建邮件对象
            msg = MIMEMultipart()
            # 发送人
            msg['from'] = sender
            # 有多个收件人时，以分号隔开
            msg['to'] = ';'.join(receiver)
            # 邮件主题
            msg['subject'] = mail_title
            '''读取测试报告内容'''
            with open(report_file, 'rb') as rf:
                mail_body = rf.read()
            '''写邮件的正文'''
            # 创建Html格式的消息对象
            body = MIMEText(mail_body, 'html', 'utf8')
            msg.attach(body)
            '''写邮件的附件'''
            # 以base64编码格式对附件编码
            att = MIMEText(mail_body, 'base64', 'utf8')
            att['Content-Type'] = 'application/octet-stream'
            att['Content-Disposition'] = 'attachment;filename = "%s"' % report_file
            msg.attach(att)
            '''发送邮件'''
            # 创建Smtp对象
            smtp = smtplib.SMTP()
            smtp.connect(smtpserver, port)
            smtp.login(sender, pwd)
            smtp.sendmail(sender, receiver.split(';'), msg.as_string())
            smtp.close()
            print('发送邮件成功')

        except:
            print('邮件发送失败')


class PathHelper:
    def get_path(self, file_path):
        '''
        获取文件的绝对路径
        :param file_path: 项目名称开始，往后的文件路径
        :return: 返回绝对路径
        '''
        project_name = file_path.split('\\')[0]
        all_path = os.path.dirname(__file__)
        path_before = all_path.split(project_name)[0]
        all_path = path_before + file_path
        return all_path
